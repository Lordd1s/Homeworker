import openpyxl

workbook_1 = openpyxl.load_workbook("hw (1)/Sheet1.xlsx")
worksheet_1 = workbook_1.active
workbook_2 = openpyxl.load_workbook("hw (1)/Sheet2.xlsx")
worksheet_2 = workbook_2.active
workbook_3 = openpyxl.load_workbook("hw (1)/Sheet3.xlsx")
worksheet_3 = workbook_3.active
workbook_main = openpyxl.Workbook()
ws_main = workbook_main.active
ws_main.sheet_properties.tabColor = "1072BA"
rows = []
ws_1 = workbook_main.create_sheet('Sheet2')
ws_2 = workbook_main.create_sheet('Sheet3')
ws_3 = workbook_main.create_sheet('Sheet4')


# TODO Decorator
def dec_up(function: any) -> any:
    def wrapper(*args, **kwargs):
        print('s')
        res = function(*args, **kwargs)
        res = str(res).capitalize()
        return res

    return wrapper


@dec_up
def word(name: str) -> str:
    return name


print(word('dias'))

# TODO First Sheet

for row_1 in range(1, worksheet_1.max_row + 1):
    local_row_1 = []
    for column in range(1, worksheet_1.max_column + 1):
        value = worksheet_1.cell(row_1, column).value
        local_row_1.append(value)
    rows.append(local_row_1)

# TODO Write to Sheet #1

for row__i, row_ in enumerate(rows, 1):
    for column__i, column_ in enumerate(row_, 1):
        ws_1.cell(row__i, column__i, value)

# TODO Second Sheet

for row_2 in range(1, worksheet_2.max_row + 1):
    local_row_2 = []
    for column in range(1, worksheet_2.max_column + 1):
        value = worksheet_2.cell(row_2, column).value
        local_row_2.append(value)
    rows.append(local_row_2)

# TODO Write to Sheet #2

l_main = []
for row_ in range(1, worksheet_2.max_row + 1):
    l1 = []
    value1 = worksheet_2.cell(row_, column=2).value
    l1.append(value1)
    l_main.append(l1)
for row__i, row_ in enumerate(l_main, 1):
    for column__i, column_ in enumerate(row_, 1):
        ws_2.cell(row__i, column__i, value)

# TODO  Third Sheet

for row_3 in range(1, worksheet_3.max_row + 1):
    local_row_3 = []
    for column in range(1, worksheet_3.max_column + 1):
        value = worksheet_3.cell(row_3, column).value
        local_row_3.append(value)
    rows.append(local_row_3)

# TODO Write to Sheet #3

l_main_2 = []
for row_ in range(1, worksheet_3.max_row + 1):
    l1_2 = []
    value1 = worksheet_3.cell(row_, column=3).value
    l1_2.append(value1)
    l_main_2.append(l1_2)
for row__i, row_ in enumerate(l_main_2, 1):
    for column__i, column_ in enumerate(row_, 1):
        ws_3.cell(row__i, column__i, value)

# TODO Write Sheet To Main

for row_i, row in enumerate(rows, 1):
    for column_i, value in enumerate(row, 1):
        ws_main.cell(row_i, column_i, value)

# TODO Save

ws_main.move_range("B20:B41", rows=-19, cols=0)
ws_main.move_range("C42:C65", rows=-41, cols=0)
workbook_main.save("EXCEL_NEWSHEEET.xlsx")
