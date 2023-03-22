import openpyxl

wb = openpyxl.load_workbook("Book1.xlsx")

wb_list = [sheet for sheet in wb.worksheets]
values = []

for one_shit in wb_list:
    for row in range(1, one_shit.max_row + 1):
        for column in range(1, one_shit.max_column + 1):
            values.append(one_shit.cell(row=row, column=column).value)


new_wb = openpyxl.Workbook()
new_ws = new_wb.active
for row, item in enumerate(values, 1):
    new_ws.cell(row=row, column=1, value=item)
new_wb.save('new.xlsx')
