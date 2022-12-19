import openpyxl

workbook = openpyxl.load_workbook(filename='Excel.xlsx')
worksheet = workbook.active

rows = []

for row in range(1, worksheet.max_row + 1, 1):
    local_row = []
    for column in range(1, worksheet.max_column + 1, 1):
        value = worksheet.cell(row, column).value
        local_row.append(value)
    rows.append(local_row)

new_workbook = openpyxl.Workbook()
new_worksheet = new_workbook.active

for row_i, row in enumerate(rows, 1):
    for column_i, value in enumerate(row, 1):
        new_worksheet.cell(column_i, row_i, value)

new_workbook.save("Excel2.xlsx")
