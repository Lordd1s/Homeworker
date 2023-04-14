import requests
import json
import os
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# response = requests.get(url="https://jsonplaceholder.typicode.com/todos").json()
# for i in response:
#     with open(f"temp/new_json{i['id']}.json", mode="w") as file:
#         json.dump(i, file)

wb = openpyxl.Workbook()
ws = wb.active

with open("temp/new_json1.json", mode="r") as file:
    data = json.load(file)
    headers = [list(data)[0], list(data)[1], list(data)[2], list(data)[3]]
    ws.append(headers)

json_files = [x for x in os.listdir('temp')]
for json_file in json_files:
    with open(f"temp/{json_file}", mode="r") as file:
        data = json.load(file)

        # print(headers)
        new_row = [data["userId"], data["id"], data['title'], data['completed']]
        row = ws.max_row+1
        for col, value in enumerate(new_row, start=1):
            ws.cell(row=row, column=col, value=value)


wb.save("excel/data.xlsx")
