import openpyxl

wb = openpyxl.load_workbook("Excel.xlsx")
ws = wb.active
ex_mas = []
max_row = ws.max_row
max_column = ws.max_column

for row in range(1, max_row + 1):
    in_mas = []
    for column in range(1, max_column + 1):
        in_mas.append(ws.cell(row, column).value)
    ex_mas.append(in_mas)
