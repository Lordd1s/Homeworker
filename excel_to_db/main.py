import openpyxl
import psycopg2


class Excel:
    def __init__(self, name_of_book: str):
        self.name_of_book = name_of_book

    def open_excel(self):
        excel_book = openpyxl.load_workbook(filename=self.name_of_book)
        excel_sheet = excel_book.active
        return excel_sheet


class DbConnectAdd:
    @staticmethod
    def add_to_db(title: str, description: str, success: bool, deadline: str, data_created: str):
        with psycopg2.connect(user="pgs_user", password="Dias15", host="127.0.0.1", port="5432", dbname="pgs_db") as connection:
            with connection.cursor() as cursor:
                try:
                    cursor.execute("INSERT INTO public.excel(title, description, success, deadline, data_created) VALUES" f"('{title}', '{description}', '{success}', '{deadline}', '{data_created}')")
                except Exception as error:
                    print(error)
                    connection.rollback()
                else:
                    connection.commit()


excel_file = Excel("hw/data.xlsx").open_excel()
rows = []

for row in range(1, excel_file.max_row + 1, 1):
    local_row = []
    for column in range(1, excel_file.max_column + 1, 1):
        value = excel_file.cell(row, column).value
        local_row.append(value)
    rows.append(local_row)

for item in rows[1::]:
    try:
        DbConnectAdd.add_to_db(title=item[1], description=item[2], success=item[3], deadline=item[4], data_created=item[5])
    except Exception as error:
        print(error)


